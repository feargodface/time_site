from django.shortcuts import render, redirect, get_object_or_404
from time_site.forms import WorkLogForm
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from time_site.models import Department, WorkLog, LeaveRequest, TeamTask
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


PAID_LEAVE_TYPES = {'отпуск', 'командировка', 'больничный', 'удалёнка'}
UNPAID_LEAVE_TYPES = {'отгул'}


@login_required
def my_statistics_view(request):
    today = now().date()
    logs = WorkLog.objects.filter(user=request.user).order_by('-date')[:30]
    if request.method == 'POST':
        form = WorkLogForm(request.POST)
        if form.is_valid():
            log, created = WorkLog.objects.get_or_create(user=request.user, date=today)
            for field in form.cleaned_data:
                setattr(log, field, form.cleaned_data[field])
            log.save()
            return redirect('my_statistics')
    else:
        form = WorkLogForm()

    return render(request, 'statistics/my_stats.html', {
        'form': form,
        'logs': logs,
        'today': today,
    })


def get_department_statistics(department):
    today = now().date()
    month_start = today.replace(day=1)

    manager = department.manager
    all_employees = department.employees.all()
    regular_employees = all_employees.exclude(id=manager.id) if manager else all_employees

    stats = []

    def get_stats_for_employee(emp, role):
        logs = WorkLog.objects.filter(user=emp, date__range=(month_start, today)).order_by('-date')
        schedule = getattr(emp.department, 'schedule', None)
        log_by_date = {log.date: log for log in logs}

        approved_leaves = LeaveRequest.objects.filter(
            user=emp,
            status='одобрено',
            end_date__gte=month_start,
            start_date__lte=today
        )

        total_minutes = 0
        worked_days = 0
        late_count = 0
        absenteeism_count = 0
        current_date = max(month_start, emp.hire_date)

        while current_date <= today:
            is_working_day = True
            if schedule:
                is_working_day = schedule.is_working_day(current_date)

            if not is_working_day:
                current_date += timedelta(days=1)
                continue

            log = log_by_date.get(current_date)

            leave = next((leave for leave in approved_leaves if leave.start_date <= current_date <= leave.end_date), None)
            leave_type = leave.leave_type if leave else None

            if log:
                total_minutes += log.worked_time_minutes
                worked_days += 1

                if log.check_in:
                    if schedule and schedule.work_start:
                        allowed_lateness = (
                            datetime.combine(current_date, schedule.work_start) + timedelta(minutes=5)
                        ).time()
                    else:
                        allowed_lateness = datetime.strptime("09:05", "%H:%M").time()

                    if log.check_in > allowed_lateness:
                        late_count += 1
            else:
                if leave_type in PAID_LEAVE_TYPES:
                    worked_days += 1
                elif leave_type in UNPAID_LEAVE_TYPES:
                    pass
                else:
                    absenteeism_count += 1

            current_date += timedelta(days=1)

        avg_minutes = total_minutes // worked_days if worked_days else 0

        stat = {
            'employee': emp,
            'role': role,
            'worked_days': worked_days,
            'total_time': total_minutes / 60,
            'avg_per_day': avg_minutes / 60 if worked_days else 0,
            'late_days': late_count,
            'absenteeism': absenteeism_count,
            'last_log': logs.first().date if logs.exists() else "—"
        }

        completed_tasks_count = TeamTask.objects.filter(
            assigned_to=emp,
            status='завершена',
            department=department
        ).count()

        stat['completed_tasks'] = completed_tasks_count

        return stat

    if manager:
        stats.append(get_stats_for_employee(manager, "Руководитель"))

    for emp in regular_employees:
        stats.append(get_stats_for_employee(emp, "Сотрудник"))

    return stats, month_start


@login_required
def department_statistics_view(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.user.role != 'Руководитель' or request.user.department != department:
        return HttpResponse(status=403)

    stats, month_start = get_department_statistics(department)

    return render(request, 'statistics/department_stats.html', {
        'department': department,
        'stats': stats,
        'month': month_start.strftime('%B %Y')
    })


@login_required
def export_department_excel(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.user.role != 'Руководитель' or request.user.department_id != department.id:
        return HttpResponse(status=403)

    stats, _ = get_department_statistics(department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика отдела"

    headers = [
        "Сотрудник", "Роль", "Отработано (дней)", "Общее время (часы)",
        "Среднее за день (часы)", "Опозданий", "Послед. рабочий день", "Задач выполнено"
    ]

    ws.append(headers)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for stat in stats:
        row = [
            stat['employee'].get_full_name() or stat['employee'].username,
            stat['role'],
            stat['worked_days'],
            round(stat['total_time'], 2),
            round(stat['avg_per_day'], 2),
            stat['late_days'],
            str(stat['last_log']) + (f" | Прогулов: {stat['absenteeism']}" if stat['absenteeism'] else ""),
            stat['completed_tasks']
        ]
        ws.append(row)
        for idx, cell in enumerate(ws[ws.max_row]):
            if idx == 0:
                cell.fill = gray_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=department_{department.pk}_stats.xlsx'
    wb.save(response)
    return response
