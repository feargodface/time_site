from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from time_site.forms import LeaveRequestForm
from time_site.models import LeaveRequest
from django.utils import translation
from django.http import HttpResponseForbidden
from time_site.models import CustomUser
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils.timezone import now
from django.http import HttpResponse


translation.activate("ru")


@login_required
def leave_form_view(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.user = request.user
            leave.save()
            return redirect('leave_list')
    else:
        form = LeaveRequestForm()
    return render(request, 'leave/leave_form.html', {'form': form})


@login_required
def leave_list_view(request):
    leaves = LeaveRequest.objects.filter(user=request.user)
    return render(request, 'leave/leave_list.html', {'leaves': leaves})


@login_required
def leave_edit_view(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk, user=request.user)
    if leave.status != 'на рассмотрении':
        return redirect('leave_list')

    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, instance=leave)
        if form.is_valid():
            form.save()
            return redirect('leave_list')
    else:
        form = LeaveRequestForm(instance=leave)

    return render(request, 'leave/leave_form.html', {'form': form, 'edit': True})


@login_required
def leave_delete_view(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk)

    if request.user != leave.user and request.user.role != 'Руководитель':
        return HttpResponseForbidden("Нет доступа для удаления заявки")

    if request.method == "POST":
        leave.delete()
        return redirect('leave_list')

    return HttpResponseForbidden("Неправильный метод запроса")


@login_required
def leave_approval_list_view(request):
    if not request.user.is_manager:
        return HttpResponseForbidden("Доступ запрещён")

    employees = CustomUser.objects.filter(department=request.user.department)
    leaves = LeaveRequest.objects.filter(user__in=employees).order_by('-submitted_at')

    return render(request, 'leave/leave_approval_list.html', {'leaves': leaves})


@login_required
def approve_leave_view(request, pk):
    if not request.user.is_manager:
        return HttpResponseForbidden()

    leave = get_object_or_404(LeaveRequest, pk=pk)
    if leave.user.department != request.user.department:
        return HttpResponseForbidden("Можно управлять только заявками своего отдела")

    leave.status = 'одобрено'
    leave.save()
    return redirect('leave_approval_list')


@login_required
def reject_leave_view(request, pk):
    if not request.user.is_manager:
        return HttpResponseForbidden()

    leave = get_object_or_404(LeaveRequest, pk=pk)
    if leave.user.department != request.user.department:
        return HttpResponseForbidden("Можно управлять только заявками своего отдела")

    leave.status = 'отклонено'
    leave.save()
    return redirect('leave_approval_list')


@login_required
def export_department_leaves_excel(request):
    if not request.user.is_manager:
        return HttpResponseForbidden("Доступ запрещён")

    employees = CustomUser.objects.filter(department=request.user.department)
    leaves = LeaveRequest.objects.filter(user__in=employees).order_by('-submitted_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки отдела"

    headers = ['Сотрудник', 'Тип', 'Дата начала', 'Дата окончания', 'Причина', 'Статус', 'Дата подачи']
    ws.append(headers)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Стили для заголовков
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Заполняем данные
    for leave in leaves:
        row = [
            leave.user.get_full_name() or leave.user.username,
            leave.get_leave_type_display(),
            leave.start_date.strftime('%d.%m.%Y'),
            leave.end_date.strftime('%d.%m.%Y'),
            leave.reason or '',
            leave.get_status_display(),
            leave.submitted_at.strftime('%d.%m.%Y %H:%M'),
        ]
        ws.append(row)

        # Стили для данных (с серым фоном в первой колонке)
        for idx, cell in enumerate(ws[ws.max_row], start=0):
            if idx == 0:
                cell.fill = gray_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # Фиксированная ширина колонок, как в функции export_department_excel
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    date_str = now().strftime("%Y%m%d")
    filename = f"заявки_отдела_{date_str}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


